"""Shared CSV / XLSX parsing utilities.

Used by application import and watchlist import. Handles:
- File-size ceiling before allocating memory
- CSV encoding (UTF-8 with BOM stripping)
- XLSX magic-byte detection so a renamed .xlsx doesn't slip through
- Header normalization (lower-cased, whitespace-trimmed, common aliases)
"""

from __future__ import annotations

import csv
import io
from collections.abc import Iterable

DEFAULT_MAX_FILE_SIZE = 2 * 1024 * 1024  # 2 MB
DEFAULT_MAX_PREVIEW_ROWS = 100

# Magic bytes so we don't trust the client's Content-Type or filename
CSV_TEXT_MAX_PROBE = 4096
XLSX_MAGIC = b"PK\x03\x04"  # xlsx is a zip container


class ImportParseError(ValueError):
    """Raised when a table file can't be parsed at all."""


def detect_format(file_obj, filename: str = "") -> str:
    """Return 'xlsx' or 'csv'.

    Reads the first few bytes to decide. `filename` is a soft hint;
    it only matters when magic bytes are ambiguous.
    """
    pos = file_obj.tell()
    try:
        head = file_obj.read(len(XLSX_MAGIC))
    finally:
        file_obj.seek(pos)

    if head == XLSX_MAGIC:
        return "xlsx"
    # Anything else is treated as CSV. We'll validate UTF-8 on parse.
    return "csv"


def _normalize_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "_")


def _read_csv(file_obj, max_size: int) -> tuple[list[str], list[dict]]:
    raw = file_obj.read(max_size + 1)
    if len(raw) > max_size:
        raise ImportParseError(f"CSV exceeds size limit ({max_size // 1024 // 1024} MB)")
    try:
        text = raw.decode("utf-8-sig")  # tolerate BOM
    except UnicodeDecodeError as e:
        raise ImportParseError("File must be UTF-8 encoded") from e

    reader = csv.DictReader(io.StringIO(text))
    headers = [_normalize_header(h) for h in (reader.fieldnames or [])]

    rows: list[dict] = []
    for row in reader:
        # Normalize keys to match `headers`
        rows.append({_normalize_header(k): (v or "").strip() for k, v in row.items()})
    return headers, rows


def _read_xlsx(file_obj, max_size: int) -> tuple[list[str], list[dict]]:
    # openpyxl doesn't stream from a file pointer with a size cap out of the
    # box, so gate on file size first.
    pos = file_obj.tell()
    file_obj.seek(0, io.SEEK_END)
    size = file_obj.tell()
    file_obj.seek(pos)
    if size > max_size:
        raise ImportParseError(f"XLSX exceeds size limit ({max_size // 1024 // 1024} MB)")

    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        raise ImportParseError(
            "Excel support is not installed on this server (openpyxl missing)."
        ) from e

    try:
        wb = load_workbook(file_obj, read_only=True, data_only=True)
    except (OSError, ValueError, KeyError) as e:
        # openpyxl raises a mix depending on the failure mode — corrupt zip
        # (BadZipFile is OSError subclass), wrong format (ValueError), etc.
        raise ImportParseError(f"Could not read XLSX file: {e}") from e

    ws = wb.active
    if ws is None:
        raise ImportParseError("XLSX has no active sheet")

    iterator: Iterable = ws.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration:
        return [], []

    headers = [_normalize_header(str(h) if h is not None else "") for h in header_row]

    rows: list[dict] = []
    for values in iterator:
        row: dict[str, str] = {}
        for h, v in zip(headers, values, strict=False):
            if not h:
                continue
            if v is None:
                row[h] = ""
            else:
                # Strip whitespace on strings; stringify others
                row[h] = str(v).strip()
        # Skip completely blank rows
        if any(row.values()):
            rows.append(row)

    wb.close()
    return headers, rows


def parse_upload(
    file_obj,
    *,
    filename: str = "",
    max_size: int = DEFAULT_MAX_FILE_SIZE,
) -> tuple[list[str], list[dict]]:
    """Parse an uploaded CSV or XLSX file into (headers, rows).

    Raises `ImportParseError` on any failure.
    """
    fmt = detect_format(file_obj, filename)
    if fmt == "xlsx":
        return _read_xlsx(file_obj, max_size)
    return _read_csv(file_obj, max_size)


def pick(row: dict, *keys: str) -> str:
    """Return the first non-empty value among the given header aliases."""
    for k in keys:
        v = row.get(k, "")
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""
